from linked_list import LinkedList # My data structure
from openpyxl import Workbook, load_workbook # For loading, creating, saving workbooks
from openpyxl.worksheet.table import Table, TableStyleInfo # For creating and styling tables
import csv # For reading csv files
import os # For file checking

filename = "sample_statement.csv"

fields = LinkedList()
rows = LinkedList()

with open(filename, "r", encoding="utf-8") as csvfile:
    csvreader = csv.reader(csvfile)
    for field in next(csvreader): # Extract fields
        if field is not None:
            fields.append(field)
    for row in csvreader: # Extract rows
        if any(row):
            row_linked_list = LinkedList()
            for field in row:
                row_linked_list.append(field)

            rows.append(row_linked_list)

    print("Total no. of rows: %d" % (csvreader.line_num))

print(f"LinkedList fields has {len(fields)} elements")
print(f"LinkedList rows has {len(rows)} elements")
print(f"The 3rd element of the 3rd rows is {rows[2][2]}")
# print('\nFirst 5 rows are:\n')
# for row in rows[:5]:
#     print(row)

print("\nVerifying rows linked list contains linked lists:\n")
for i, row_linked_list in enumerate(rows, start=1):
    # Check if the element is a LinkedList
    if isinstance(row_linked_list, LinkedList):
        print(f"Row {i} is a LinkedList. Contents: {[value for value in row_linked_list]}")
    else:
        print(f"Row {i} is NOT a LinkedList. Found type: {type(row_linked_list)}")

if os.path.exists("budget.xlsx"): # If workbook exist, then edit it
    workbook = load_workbook("budget.xlsx")
    
    if "budget" not in workbook.sheetnames:
        worksheet = workbook.create_sheet("budget")
    else:
        worksheet = workbook ["budget"]

else: # If workbook doesnt exist, then create it
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "budget"

    # Header names
    worksheet["A1"] = "Year"
    worksheet["B1"] = "Month"
    worksheet["C1"] = "Date"
    worksheet["D1"] = "Description"
    worksheet["E1"] = "Category"
    worksheet["F1"] = "Income"
    worksheet["G1"] = "Debits"
    worksheet["H1"] = "Balance"
    worksheet["I1"] = "Essential"

    # Fill the formulas
    for row_index in range(2, len(rows) + 2):
        worksheet[f"A{row_index}"] = f'=YEAR(C{row_index})'
        worksheet[f"B{row_index}"] = f'=TEXT(DATE(2011,MONTH(C{row_index}),1),"MMMM")'
        worksheet

table_range = f"A1:I{len(rows) + 1}"
table = Table(displayName = "BudgetTable", ref = table_range)

style = TableStyleInfo(
    name="TableStyleMedium4",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True,
)
table.tableStyleInfo = style

worksheet.add_table(table)

workbook.save("budget.xlsx") # Finish editing the workbook
workbook.close()


